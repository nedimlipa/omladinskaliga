"""Mini Rukomet — admin upravljanje i javni prikaz."""
import datetime
import io
from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from ..templates_config import templates, local_dt_str
from ..database import get_db
from ..models import MiniRukometTurnir, MiniRukometUtakmica, MiniRukometPrijava, Klub
from .auth import get_current_user

router = APIRouter()

SORT_OPCIJE = {
    "bodovi":          "Bodovi",
    "gol_razlika":     "Gol razlika",
    "gol_postignuti":  "Postignuti golovi",
    "gol_primljeni":   "Primljeni golovi",
    "pobjede":         "Pobjede",
    "porazi":          "Porazi",
    "utakmice":        "Odigrane utakmice",
}

# ─────────────────────────────────────────────────────────────
#  Pomoćna funkcija: izračunaj tabelu iz utakmica
# ─────────────────────────────────────────────────────────────

def _izracunaj_tabelu(utakmice: list, sort1: str, sort2: str, sort3: str) -> list:
    """Vrati sortirani niz redova tabele računajući iz utakmica."""
    ekipe: dict[str, dict] = {}

    def get_or_create(naziv: str) -> dict:
        if naziv not in ekipe:
            ekipe[naziv] = {
                "ekipa": naziv,
                "utakmice": 0, "pobjede": 0, "remiji": 0, "porazi": 0,
                "gol_postignuti": 0, "gol_primljeni": 0,
                "gol_razlika": 0, "bodovi": 0,
            }
        return ekipe[naziv]

    for u in utakmice:
        if u.gol_a is None or u.gol_b is None:
            continue  # utakmica bez rezultata — preskoči
        a = get_or_create(u.ekipa_a)
        b = get_or_create(u.ekipa_b)

        a["utakmice"] += 1
        b["utakmice"] += 1
        a["gol_postignuti"] += u.gol_a
        a["gol_primljeni"]  += u.gol_b
        b["gol_postignuti"] += u.gol_b
        b["gol_primljeni"]  += u.gol_a

        if u.gol_a > u.gol_b:
            a["pobjede"] += 1; a["bodovi"] += 3
            b["porazi"]  += 1
        elif u.gol_a < u.gol_b:
            b["pobjede"] += 1; b["bodovi"] += 3
            a["porazi"]  += 1
        else:
            a["remiji"] += 1; a["bodovi"] += 1
            b["remiji"] += 1; b["bodovi"] += 1

    for e in ekipe.values():
        e["gol_razlika"] = e["gol_postignuti"] - e["gol_primljeni"]

    def sort_key(e):
        def val(k):
            return e.get(k, 0)
        return (-val(sort1), -val(sort2), -val(sort3))

    return sorted(ekipe.values(), key=sort_key)


# ─────────────────────────────────────────────────────────────
#  JAVNI PRIKAZ
# ─────────────────────────────────────────────────────────────

@router.get("/mini-rukomet", response_class=HTMLResponse)
async def public_mini_rukomet(request: Request, db: AsyncSession = Depends(get_db)):
    turniri = (await db.execute(
        select(MiniRukometTurnir)
        .where(MiniRukometTurnir.aktivan == True)
        .order_by(MiniRukometTurnir.kreiran_datum.desc())
    )).scalars().all()

    return templates.TemplateResponse("mini_rukomet.html", {
        "request": request,
        "turniri": turniri,
    })


@router.get("/mini-rukomet/{turnir_id}", response_class=HTMLResponse)
async def public_mini_rukomet_detalji(
    turnir_id: int, request: Request, db: AsyncSession = Depends(get_db)
):
    turnir = await db.get(MiniRukometTurnir, turnir_id)
    if not turnir or not turnir.aktivan:
        return RedirectResponse("/mini-rukomet", status_code=302)

    utakmice = (await db.execute(
        select(MiniRukometUtakmica)
        .where(MiniRukometUtakmica.turnir_id == turnir_id)
        .order_by(MiniRukometUtakmica.datum_utakmice.asc().nullslast(), MiniRukometUtakmica.id.asc())
    )).scalars().all()

    tabela = _izracunaj_tabelu(utakmice, turnir.sort1, turnir.sort2, turnir.sort3)

    return templates.TemplateResponse("mini_rukomet_detalji.html", {
        "request": request,
        "turnir":  turnir,
        "utakmice": utakmice,
        "tabela":  tabela,
        "sort_opcije": SORT_OPCIJE,
    })


# ─────────────────────────────────────────────────────────────
#  ADMIN — lista turnira
# ─────────────────────────────────────────────────────────────

@router.get("/admin/mini-rukomet", response_class=HTMLResponse)
async def admin_mr_lista(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    turniri = (await db.execute(
        select(MiniRukometTurnir).order_by(MiniRukometTurnir.kreiran_datum.desc())
    )).scalars().all()

    return templates.TemplateResponse("admin_mini_rukomet.html", {
        "request": request,
        "user":    user,
        "turniri": turniri,
        "sort_opcije": SORT_OPCIJE,
        "ok":    request.query_params.get("ok"),
        "error": request.query_params.get("error"),
    })


@router.post("/admin/mini-rukomet/novi")
async def admin_mr_novi(
    request: Request,
    naziv:  str = Form(...),
    opis:   str = Form(""),
    sort1:  str = Form("bodovi"),
    sort2:  str = Form("gol_razlika"),
    sort3:  str = Form("gol_postignuti"),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    db.add(MiniRukometTurnir(
        naziv=naziv.strip(),
        opis=opis.strip() or None,
        sort1=sort1, sort2=sort2, sort3=sort3,
    ))
    await db.commit()
    return RedirectResponse("/admin/mini-rukomet?ok=1", status_code=302)


@router.post("/admin/mini-rukomet/{tid}/edit")
async def admin_mr_edit(
    tid: int,
    request: Request,
    naziv:  str  = Form(...),
    opis:   str  = Form(""),
    sort1:  str  = Form("bodovi"),
    sort2:  str  = Form("gol_razlika"),
    sort3:  str  = Form("gol_postignuti"),
    aktivan: str = Form("on"),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    t = await db.get(MiniRukometTurnir, tid)
    if t:
        t.naziv   = naziv.strip()
        t.opis    = opis.strip() or None
        t.sort1   = sort1
        t.sort2   = sort2
        t.sort3   = sort3
        t.aktivan = (aktivan == "on")
        await db.commit()
    return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok=1", status_code=302)


@router.post("/admin/mini-rukomet/{tid}/delete")
async def admin_mr_delete(
    tid: int, request: Request, db: AsyncSession = Depends(get_db)
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    t = await db.get(MiniRukometTurnir, tid)
    if t:
        await db.delete(t)
        await db.commit()
    return RedirectResponse("/admin/mini-rukomet?ok=2", status_code=302)


# ─────────────────────────────────────────────────────────────
#  ADMIN — detalji turnira (utakmice)
# ─────────────────────────────────────────────────────────────

@router.get("/admin/mini-rukomet/{tid}", response_class=HTMLResponse)
async def admin_mr_detalji(
    tid: int, request: Request, db: AsyncSession = Depends(get_db)
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    turnir = await db.get(MiniRukometTurnir, tid)
    if not turnir:
        return RedirectResponse("/admin/mini-rukomet", status_code=302)

    utakmice = (await db.execute(
        select(MiniRukometUtakmica)
        .where(MiniRukometUtakmica.turnir_id == tid)
        .order_by(MiniRukometUtakmica.datum_utakmice.asc().nullslast(), MiniRukometUtakmica.id.asc())
    )).scalars().all()

    tabela = _izracunaj_tabelu(utakmice, turnir.sort1, turnir.sort2, turnir.sort3)

    # Dohvati listu svih timova koji su ikad igrali (za autocomplete)
    ekipe_set: set[str] = set()
    for u in utakmice:
        ekipe_set.add(u.ekipa_a)
        ekipe_set.add(u.ekipa_b)

    # Prijave klubova na ovaj turnir
    prijave_rows = (await db.execute(
        select(MiniRukometPrijava, Klub)
        .join(Klub, MiniRukometPrijava.klub_id == Klub.id)
        .where(MiniRukometPrijava.turnir_id == tid)
        .order_by(MiniRukometPrijava.kreiran_datum.desc())
    )).all()
    prijave = [{"p": p, "klub": k} for p, k in prijave_rows]

    return templates.TemplateResponse("admin_mini_rukomet_detalji.html", {
        "request":  request,
        "user":     user,
        "turnir":   turnir,
        "utakmice": utakmice,
        "tabela":   tabela,
        "ekipe":    sorted(ekipe_set),
        "sort_opcije": SORT_OPCIJE,
        "prijave":  prijave,
        "ok":    request.query_params.get("ok"),
        "error": request.query_params.get("error"),
    })


# ─────────────────────────────────────────────────────────────
#  ADMIN — dodaj / izmijeni / obriši utakmicu
# ─────────────────────────────────────────────────────────────

def _parse_dt(datum: str, vrijeme: str) -> datetime.datetime | None:
    """Spoji datum i vrijeme u tz-aware datetime (Sarajevo → UTC)."""
    from zoneinfo import ZoneInfo
    _TZ = ZoneInfo("Europe/Sarajevo")
    d = datum.strip()
    v = (vrijeme.strip() or "00:00")
    if not d:
        return None
    try:
        local = datetime.datetime.strptime(f"{d} {v}", "%Y-%m-%d %H:%M")
        return local.replace(tzinfo=_TZ).astimezone(datetime.timezone.utc)
    except ValueError:
        return None


@router.post("/admin/mini-rukomet/{tid}/utakmica/novi")
async def admin_mr_utakmica_novi(
    tid: int,
    request: Request,
    datum:   str = Form(""),
    vrijeme: str = Form(""),
    ekipa_a: str = Form(...),
    ekipa_b: str = Form(...),
    gol_a:   str = Form(""),
    gol_b:   str = Form(""),
    kolo:    str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    db.add(MiniRukometUtakmica(
        turnir_id=tid,
        datum_utakmice=_parse_dt(datum, vrijeme),
        ekipa_a=ekipa_a.strip(),
        ekipa_b=ekipa_b.strip(),
        gol_a=int(gol_a) if gol_a.strip().isdigit() else None,
        gol_b=int(gol_b) if gol_b.strip().isdigit() else None,
        kolo=int(kolo) if kolo.strip().isdigit() else None,
    ))
    await db.commit()
    return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok=1", status_code=302)


@router.post("/admin/mini-rukomet/{tid}/utakmica/{uid}/edit")
async def admin_mr_utakmica_edit(
    tid: int, uid: int,
    request: Request,
    datum:   str = Form(""),
    vrijeme: str = Form(""),
    ekipa_a: str = Form(...),
    ekipa_b: str = Form(...),
    gol_a:   str = Form(""),
    gol_b:   str = Form(""),
    kolo:    str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    u = await db.get(MiniRukometUtakmica, uid)
    if u and u.turnir_id == tid:
        u.datum_utakmice = _parse_dt(datum, vrijeme)
        u.ekipa_a = ekipa_a.strip()
        u.ekipa_b = ekipa_b.strip()
        u.gol_a   = int(gol_a) if gol_a.strip().isdigit() else None
        u.gol_b   = int(gol_b) if gol_b.strip().isdigit() else None
        u.kolo    = int(kolo)  if kolo.strip().isdigit()  else None
        await db.commit()
    return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok=1", status_code=302)


@router.post("/admin/mini-rukomet/{tid}/utakmica/{uid}/delete")
async def admin_mr_utakmica_delete(
    tid: int, uid: int, request: Request, db: AsyncSession = Depends(get_db)
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    u = await db.get(MiniRukometUtakmica, uid)
    if u and u.turnir_id == tid:
        await db.delete(u)
        await db.commit()
    return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok=1", status_code=302)


# ─────────────────────────────────────────────────────────────
#  ADMIN — uvoz iz Excela
# ─────────────────────────────────────────────────────────────

@router.post("/admin/mini-rukomet/{tid}/upload-excel")
async def admin_mr_upload_excel(
    tid: int,
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)

    turnir = await db.get(MiniRukometTurnir, tid)
    if not turnir:
        return RedirectResponse("/admin/mini-rukomet", status_code=302)

    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # Pronađi header red (Datum, Vrijeme, Ekipa A, Ekipa B, Rezultat)
        header_map: dict[str, int] = {}
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
            if "datum" in row_lower or "ekipa a" in row_lower:
                header_row = i
                for j, h in enumerate(row_lower):
                    header_map[h] = j
                break

        if header_row is None:
            return RedirectResponse(f"/admin/mini-rukomet/{tid}?error=header", status_code=302)

        # Kolone po imenu (fleksibilno)
        ci_datum   = header_map.get("datum",   header_map.get("date",   None))
        ci_vrijeme = header_map.get("vrijeme", header_map.get("time",   header_map.get("vr", None)))
        ci_a       = header_map.get("ekipa a", header_map.get("ekipa_a", header_map.get("home", None)))
        ci_b       = header_map.get("ekipa b", header_map.get("ekipa_b", header_map.get("away", None)))
        ci_rez     = header_map.get("rezultat", header_map.get("result", header_map.get("score", None)))
        ci_kolo    = header_map.get("kolo",    header_map.get("round",  header_map.get("runda", None)))

        if ci_a is None or ci_b is None:
            return RedirectResponse(f"/admin/mini-rukomet/{tid}?error=cols", status_code=302)

        inserted = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(c is None for c in row):
                continue  # prazni redovi

            def cell(ci):
                return str(row[ci]).strip() if ci is not None and ci < len(row) and row[ci] is not None else ""

            ea = cell(ci_a)
            eb = cell(ci_b)
            if not ea or not eb:
                continue

            # Parsiraj rezultat (npr. "3:2" ili "3-2")
            gol_a = gol_b = None
            rez_str = cell(ci_rez)
            if rez_str:
                import re
                m = re.match(r"^(\d+)\s*[:–\-]\s*(\d+)$", rez_str)
                if m:
                    gol_a, gol_b = int(m.group(1)), int(m.group(2))

            # Parsiraj datum
            datum_str   = cell(ci_datum)
            vrijeme_str = cell(ci_vrijeme)
            dt = None
            if datum_str and datum_str not in ("None", ""):
                # Probaj različite formate
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        datum_obj = datetime.datetime.strptime(datum_str, fmt)
                        # Spoji s vremenom
                        vr_parsed = "00:00"
                        if vrijeme_str and vrijeme_str not in ("None", ""):
                            for vfmt in ("%H:%M", "%H:%M:%S", "%H.%M"):
                                try:
                                    vr_parsed = datetime.datetime.strptime(
                                        vrijeme_str.split(".")[0].strip()[:5], "%H:%M"
                                    ).strftime("%H:%M")
                                    break
                                except ValueError:
                                    continue
                        dt = _parse_dt(datum_obj.strftime("%Y-%m-%d"), vr_parsed)
                        break
                    except ValueError:
                        continue

            # Parsiraj kolo
            kolo_val = None
            kolo_str = cell(ci_kolo)
            if kolo_str and kolo_str.isdigit():
                kolo_val = int(kolo_str)

            db.add(MiniRukometUtakmica(
                turnir_id=tid,
                datum_utakmice=dt,
                ekipa_a=ea,
                ekipa_b=eb,
                gol_a=gol_a,
                gol_b=gol_b,
                kolo=kolo_val,
            ))
            inserted += 1

        await db.commit()
        return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok={inserted}", status_code=302)

    except Exception as e:
        return RedirectResponse(f"/admin/mini-rukomet/{tid}?error=parse", status_code=302)


# ─────────────────────────────────────────────────────────────
#  ADMIN — upravljanje prijavama klubova
# ─────────────────────────────────────────────────────────────

@router.post("/admin/mini-rukomet/{tid}/prijava/{pid}/odobri")
async def admin_mr_prijava_odobri(
    tid: int, pid: int, request: Request, db: AsyncSession = Depends(get_db)
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)
    prijava = await db.get(MiniRukometPrijava, pid)
    if prijava and prijava.turnir_id == tid:
        prijava.status = "odobren"
        await db.commit()
    return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok=odobreno", status_code=302)


@router.post("/admin/mini-rukomet/{tid}/prijava/{pid}/odbij")
async def admin_mr_prijava_odbij(
    tid: int, pid: int, request: Request, db: AsyncSession = Depends(get_db)
):
    user = get_current_user(request)
    if not user or user.get("tip") not in ("admin", "moderator"):
        return RedirectResponse("/login", status_code=302)
    prijava = await db.get(MiniRukometPrijava, pid)
    if prijava and prijava.turnir_id == tid:
        await db.delete(prijava)
        await db.commit()
    return RedirectResponse(f"/admin/mini-rukomet/{tid}?ok=odbijeno", status_code=302)


# ─────────────────────────────────────────────────────────────
#  KLUB — prijava na mini rukomet turnir
# ─────────────────────────────────────────────────────────────

@router.post("/klub/mini-rukomet/prijava")
async def klub_mr_prijava(
    request: Request,
    turnir_id:   int = Form(...),
    naziv_ekipe: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user or user.get("tip") != "klub":
        return RedirectResponse("/login", status_code=302)
    klub_id = int(user["sub"])

    naziv_ekipe = naziv_ekipe.strip()
    if not naziv_ekipe:
        return RedirectResponse("/klub/dashboard?error=mr_naziv", status_code=302)

    # Provjeri da li već postoji prijava za ovaj turnir
    existing = (await db.execute(
        select(MiniRukometPrijava).where(
            MiniRukometPrijava.turnir_id == turnir_id,
            MiniRukometPrijava.klub_id == klub_id,
        )
    )).scalar_one_or_none()

    if not existing:
        turnir = await db.get(MiniRukometTurnir, turnir_id)
        if turnir and turnir.aktivan:
            db.add(MiniRukometPrijava(
                turnir_id=turnir_id,
                klub_id=klub_id,
                naziv_ekipe=naziv_ekipe,
            ))
            await db.commit()

    return RedirectResponse("/klub/dashboard?ok=mr_prijava", status_code=302)


@router.post("/klub/mini-rukomet/prijava/{pid}/obrisi")
async def klub_mr_prijava_obrisi(
    pid: int, request: Request, db: AsyncSession = Depends(get_db)
):
    user = get_current_user(request)
    if not user or user.get("tip") != "klub":
        return RedirectResponse("/login", status_code=302)
    klub_id = int(user["sub"])
    prijava = await db.get(MiniRukometPrijava, pid)
    if prijava and prijava.klub_id == klub_id and prijava.status == "na_cekanju":
        await db.delete(prijava)
        await db.commit()
    return RedirectResponse("/klub/dashboard", status_code=302)

